
import os
from gtts import gTTS
import speech_recognition as sr
from math import remainder


para = '''
Calculate the mean score of all users.                                            # sum([ i['age'] for i in Student.objects.all().values('age')  ])/2   
Find the user with the highest score.                                             # Student.objects.order_by('-age')[:1].values('age')[0]    
Find if there is any user with score less than 2 (return True/False)              #  User.objects.filter(age__lt=20).exists()           |  [True if Student.objects.filter(age__lt=20) else False]     | 
Find all the users which have score exactly equal to the mean score.              # Student.objects.filter(age=sum([ i['age'] for i in Student.objects.all().values('age') ])//2)
Change the score value of all users in a single query to 100 (one hundred)        #
Find all the users which don't have score 10 or 20                                # Student.objects.exclude(age__in=[10,20,31])
Print a list of all the score values from all users excluding the first 10 users. # Student.objects.exclude(id__in=[i for i in range(10)]).values('age')
'''
# d={}
# words=para.split()
# for i in words:
#      if i not in d:
#           d[i] =0
#      else:
#           d[i] += 1

# # print(d)


# L=[1,2,4,5,0,6,0]
# def solve(L):
#      if len(L) == 0:
#           return L
#      k=0
#      for i in range(
#
# len(L)):
#           if L[i] !=0:
#                L[k] = L[i]
#                k+=1
#      for o in range(k,len(L)):
#           L[o] =0
#           return L
# print(solve(L))


# print(    sorted(d.items() , key = lambda kv:(kv[1],kv[0])) )


# def hi(func):
#      print('hello this is the decorator ')
#      def wrapper(a):
#           print(f' calling {a} from inside the wrapper')
#           # return a+'hi'
#           return a   # this  refers to line 58 , where it gets called accordingly
#      return wrapper

# @hi
# def call(val):
#      print(val+'malhan')

# call('tushar')

# def input_name(func):
#      def wrapper(name):
#           name = input('Enter your name: ')
#           if name:
#                print(func(name))        # this will call decorated function,otherwise it has not meaning
#                return 'hi ' + name      
#           else:
#                return 'Please enter your name'
#      return wrapper

# @input_name
# def say(name):
#      return 'hi'+name
# print(say('malhan'))

# from chatterbot import ChatBot
# from chatterbot.trainers import ListTrainer
# from chatterbot.trainers import CHatterBotCorpus Trainer

# logic_adapters = (['chatterbot.logic.Mathematical Evalution',
# 'chatterbot.logic.BestMatch'])


# smalltalk = ['hi','hai','how do you do',
# 'hello','hai','hey','hi','how are you','hey buddy','you are great','your great','hay','ok'
# ]


# talk1= ['best thing ever to take place ']

# list_trainer= List Trainer (my_bot)
# for item in (smalltalk,talk1):
#      list_trainer.train(item)


# corpus_trainer = ChatterBotCorpus Trainer(my_bot)
# corpus_trainer.train('chatterbot.corpus.english')

# print(my_bot.get_response('hello'))

# from gtts import gTTS
# voice = ''

# while True:
#      e = sr.Recognizer()
#      with sr.Microphone() as source:
#           try:
#                audio = e. listen(source)
#                text = e.recognize_google(audio)
#                print(text)
#                if text == "stop":
#                     break
#                text = e.recognize_google(audio)
#                voice = voice+str(text)

#           except:
#                 print("say something")

# we = gTTS(text=voice, Lang='en', slow=False)
# we.save("A.wav")

# from GoogleNews import GoogleNews
# googlenews = GoogleNews( )
# googlenews = GoogleNews(period='/7d' )
# googlenews.search( 'IND' )
# result=googlenews.result( )

# for x in result:
#      print("-"*50)
#      print("Title--", x['title'])
#      print("Date/Time--", x['date'])
#      print("Description--", x['desc'])

#      print("Link--", x['link'])


########################################################################
# CREATE CSV FILE AND APPEND TO IT 

# import csv

# csv_file = open('sample.csv', 'w')
# csv_writer = csv.writer(csv_file)
# csv_writer.writerow(['Headline', 'Summary', 'Video']) # headers of the file
# for i in range(10):
#      csv_writer.writerow([i])
# for i in range(10,0,-1):
#      csv_writer.writerow([i,i+1,i*2,i==i])
# csv_file.close()

################################################################################################################################################
#  WEB SCRAPING (extract  from web and save content in excel file )

from bs4 import BeautifulSoup
from openpyxl import Workbook  
import time  

# wb = Workbook()  
# sheet = wb.active 


def file_extract(filename):
     with open(filename,'r') as f:
          html=f.read()
          soup = BeautifulSoup(html, 'html.parser')
          name_tag = soup.find_all('h4',class_='subhead1-r product-cell__description-name')
          names_left = soup.find_all('button',class_='cart-product-cell__image')
          price_tag= soup.find_all('p',class_='product-price__unit-price subhead1-b')
          
          for word in range(len(name_tag)):
               name_of_the_product=name_tag[word].get_text()
               # sheet[f'A{word+1}']=name_of_the_product
         
          
          # for name in names_left:
          #      #Peanut butter 100% Hacendado - 3,00
          #      #Deep-frozen whole strawberries Hacendado -1.50
 
               # print(name['aria-label']) 
               ...
          # sheet['A61']='Peanut butter 100% Hacendado'
          # sheet['A62']='Deep-frozen whole strawberries Hacendado'

          for word in range(len(price_tag)):
               price_of_the_product=price_tag[word].get_text().replace('â‚¬','').replace(',','.')
               print(price_of_the_product)
               # sheet[f'B{word+1}']=price_of_the_product

# file_extract('part2.html')

# wb.save("excel/sheet1.xlsx")  


######################        EXTRACT DATA FROM MULTIPLE INNER HTML'S AND SAVING FILES TO EXCEL   DYNAMICALLY      #############################
#309

# from selenium import webdriver
# import time
# from selenium.webdriver.common.keys import Keys
# from bs4 import BeautifulSoup

# from selenium.webdriver.common.by import By 
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC

# from openpyxl import Workbook  
# import time  

# wb = Workbook()  
# sheet = wb.active 

"""       """                   


# from selenium import webdriver
# import time
# from selenium.webdriver.common.keys import Keys
# from bs4 import BeautifulSoup

# from selenium.webdriver.common.by import By 
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC

# from openpyxl import Workbook  
# import time  

# wb = Workbook()  

# # checks on ur param
# url="https://www.mercadona.es/"
# postalcode='08010'

# def get_values(sheet_num,sheet_name,url,postcode,class_name1,class_name2,class_name3,class_name4,class_name5,class_name6,driver):
          
#           sheet = wb.create_sheet(sheet_num)
#           sheet.title = sheet_name
          
#           # get the Name
#           c=1
#           sheet[f'A{c}']="PRODUCTS"
#           name_tag= driver.find_elements_by_class_name(class_name1)
#           for name in name_tag:
#                c+=1
#                sheet[f'A{c}']=name.text
#                print(name.text)
          

#           # get the Price
#           c2=1
#           sheet[f'B{c2}']="PRICES"
#           price_tag=driver.find_elements_by_class_name(class_name2)
#           for price in price_tag:
#                c2+=1
#                sheet[f'B{c2}']=price.text
#                print(price.text)

         
#           #                Get all the images link 
#           #                        option 1 
#           # images = driver.find_elements_by_tag_name('img')     # get all the images                                    
          
#           #                        option 2
#           images = driver.find_elements_by_xpath(class_name3)    # //div[@class='image_class_name']/img
#           c3=1
#           sheet[f'C{c3}']="IMAGES"
#           for image in images:
#                c3+=1
#                sheet[f'C{c3}']=image.get_attribute('src')
#                print(image.get_attribute('src'))
#                # alt = image.get_attribute("alt") 
#           # print(c3)

#           print('\n\nDone , Data extracted and saved in your excel ')
          
# # final execution
# def web_scrap_automation(url,postcode,class_name1=None,class_name2=None,class_name3=None,class_name4=None,class_name5=None,class_name6=None):
 
#      driver = webdriver.Chrome(executable_path=r"C:\Users\DELL\Downloads\chromedriver.exe")
#      driver.get(url)
#      print('\n\n')
#      print(driver.title)
#      print('\n\n')
#      try:
#           #              convert to english language
#           l0=driver.find_element_by_xpath("/html/body/div[1]/nav/div[2]/div/i")
#           l0.click()
#           l1=driver.find_element_by_xpath("/html/body/div[1]/nav/div[2]/div/ul/li[6]/button")
#           l1.click()
          
#           search_bar = driver.find_element_by_xpath('//*[@id="root"]/header/div/div/form/div/input')
          
#           #                   give the postcode
#           search_bar.send_keys(postcode)      
#           search_bar.send_keys(Keys.RETURN)
#           l2=driver.find_element_by_xpath('//*[@id="root"]/header/div/div/form/input')
#           l2.click()
#           accept= driver.find_element_by_xpath('//*[@id="root"]/div[1]/div/div/button[2]')
#           accept.click()

#           time.sleep(10)
#           print('\n\n')

#           #                        DYNAMIC FROM HERE 

#           get_values('sheet_1','LANDING_PAGE', url,postcode,class_name1,class_name2,class_name3,class_name4,class_name5,class_name6,driver=driver)
#           print('\n\n')
     
#           #                      View the products from inside 
#           product_button = driver.find_element_by_class_name('banner-item__button')
#           product_button.click()
#           time.sleep(10)


#           print('\n\n')
#           get_values('sheet_2','INNER_PAGE',url,postcode,class_name1,class_name2,class_name3,class_name4,class_name5,class_name6,driver=driver)

#           time.sleep(5)
#           # driver.close()                                                                 
#           # wb.save(r"C:\Users\DELL\Desktop\excel\sheet2.xlsx")  
#           wb.save(filename = fr"C:\Users\DELL\Desktop\excel\file_{postalcode}.xlsx")             # saving file to the excel 

#      except:
#           print(f'Issue in the postal code ❌{postalcode}❌, check and verify it again')
          
# web_scrap_automation(url,postalcode,'product-cell__description-name','product-price__unit-price',"//div[@class='product-cell__image-wrapper']/img")

# web_scrap_automation("https://www.amazon.in/deal/825e22a9?showVariations=true&smid=A14CZOWI0VEHLG&pf_rd_r=J4WQV8JHJ7TXSR3DVWXP&pf_rd_p=f690369a-7bb4-48bb-9349-f68a447ef06d&pd_rd_r=c55c1e7a-619a-4698-8aab-7c79a9509c58&pd_rd_w=jsird&pd_rd_wg=CHlPw&ref_=pd_gw_unk",'08012','a-text-normal','a-price-whole')


""" #############################################################################################################################################################################"""
#                                 Shorter version of above code               
#                 just provide the class name and ask driver to find element by respective type and add the functionality inside get_values()

# no big functions , one main function  |  inside class 
# no comments for prod 

from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook  ,load_workbook
wb = Workbook()  
url="https://www.mercadona.es/"
postalcode='08013'

def web_scrap_automation(url,postcode,*classnames):
     driver = webdriver.Chrome(executable_path=r"C:\Users\DELL\Downloads\chromedriver.exe")
     driver.get(url)
     print('\t\t',driver.title)
     
     def get_values(sheet_num,sheet_name,driver,*innerclasses):     # here innerclasses refers to each element we sent when outer func is called 
               sheet = wb.create_sheet(sheet_num)
               sheet.title = sheet_name 
               
               #   get the Name
               sheet[f'A{1}']="PRODUCTS"
               name_tag= driver.find_elements_by_class_name(innerclasses[0])
               for name in range(len(name_tag)):
                    sheet[f'A{name+2}']=name_tag[name].text
                    print(name_tag[name].text)

               
               #   get the Price
               sheet[f'B{1}']="PRICES"
               price_tag=driver.find_elements_by_class_name(innerclasses[1])
               for price in range(len(price_tag)):
                    sheet[f'B{price+2}']= price_tag[price].text
                    print(price_tag[price].text)

               #  getting the image link
               images = driver.find_elements_by_xpath(innerclasses[2]) 
               sheet[f'C{1}']="IMAGES"
               for image in range(len(images)):
                    sheet[f'C{image+2}']=images[image].get_attribute('src')
                    print(images[image].get_attribute('src'))

               # get the weight
               weights =driver.find_elements_by_class_name(innerclasses[3]) 
               sheet[f'D{1}']="WEIGHT"
               for weight in range(len(weights)):
                    sheet[f'D{weight+2}'] = weights[weight].text
                    print(weights[weight].text)
               print()
                    
               print(f'\n\nDone , Data extracted and saved in your excel for the {sheet_name}')

     try:
          #              convert to english language
          l0=driver.find_element_by_xpath("/html/body/div[1]/nav/div[2]/div/i")
          l0.click()
          l1=driver.find_element_by_xpath("/html/body/div[1]/nav/div[2]/div/ul/li[6]/button")
          l1.click()
          
          search_bar = driver.find_element_by_xpath('//*[@id="root"]/header/div/div/form/div/input')
          
          #                   give the postcode
          try:
               search_bar.send_keys(postcode)      
               search_bar.send_keys(Keys.RETURN)
               l2=driver.find_element_by_xpath('//*[@id="root"]/header/div/div/form/input')
               l2.click()
               accept= driver.find_element_by_xpath('//*[@id="root"]/div[1]/div/div/button[2]')
               accept.click()
          except:
               print('No way to pass postal code on the webpage')

          time.sleep(10)
          #                        DYNAMIC FROM HERE 

          get_values('sheet_1','LANDING_PAGE',driver,*classnames)
          print('\n\n')
     
          # #                      View the products from inside 
          product_button = driver.find_element_by_class_name('banner-item__button')
          product_button.click()
          time.sleep(10)

          get_values('sheet_2','INNER_PAGE',driver,*classnames)
          print('\n\n')

          driver.close()      
          filename = fr"C:\Users\DELL\Desktop\excel\file_{postalcode}.xlsx"                                                           
          wb.save(filename)                                                          # saving file to the excel 
          wb.remove(wb['Sheet'])
          wb.save(filename)  

     except Exception as e:
          print(f'Issue in the postal code ❌{postalcode}❌, check and verify it again')
          print(e)

# web_scrap_automation(url,postalcode,'product-cell__description-name','product-price__unit-price',"//div[@class='product-cell__image-wrapper']/img","product-price__extra-price",None)

""" #############################################################################################################################################################################""" 


#   HERE IF WE CAN COMMENT 2 FN CALL , WE CAN WEB SCRAPE ALMOST EVERY SITE 
######################################################################################################################################################

#    ubuntu multiple commands at once

# !wsl && cd /mnt/c/Users/DELL/Desktop/PYTHON/test/RealTimeObjectDetection/Tensorflow && git clone https://github.com/tensorflow/models

########################################################################

# save an excel file
# from openpyxl.workbook import Workbook

# POINTS 
# IF FILE IS OPENEED , THEN IT WILL NOT SAVE THE FILE , ERROR 


#              Create sheets in the workbook

# from openpyxl.workbook import Workbook
# wb = Workbook()
# ws1 = wb.create_sheet("Sheet_1")                          # if created with ws1 , dont set this as active
# ws1.title = "Title_A>@"
# ws1['A1'] = 'PRODUCTS'
# ws1['B1'] = 'PRICE'
# ws1['C1'] = 'IMAGE'


# ws2 = wb.create_sheet("Sheet_2")
# ws2.title = "Title_A>!"
# ws2['A1'] = 'PRODUCTS'
# ws2['B1'] = 'PRICE'
# ws2['C1'] = 'IMAGE'

# wb.save(filename = r"C:\Users\DELL\Desktop\excel\tem.xlsx")


#              use an excel file to append

# from openpyxl import Workbook  
# import time  
  
# wb = Workbook()  
# sheet = wb.active  


# sheet['A1'] = 87  
# sheet['A2'] = "Devansh"  
# sheet['A3'] = 41.80  
# sheet['A4'] = 10  
  
# now = time.strftime("%x")  
# sheet['A5'] = now  
  
# wb.save("excel/sheet1.xlsx")           # create new sheet if not present

########################################################################

     # Access variables from function

# global a : print(a)      
# create global variable  , use the same variable for the function , to call the global variable 
# inisde func => func.a =1  , when func called , print(func.a)

# use it to create global and local variables inisde function
# globals()             # will return all global variables 
# x = globals()['a']    # will return specific global variable outside the function which is a


########################################################################

#                    change text from file 

# os.chdir(r'C:\Users\DELL\Desktop\personal\no_specs')
# list_files=[file for file in os.listdir() if file.endswith('txt') ] 
# # print(list_files)

# for file in list_files:
#      with open(file,'r+')as f:             # first read the file 
                                             
#           line = f.read()                  # create a new line with changes
#           change=(line.replace(''.join(line.split(' ')[0]) ,'0'))

# for file in list_files:
#      with open(file,'r+')as f:
#           ...
#           f.write(change)                   # perform

########                                      ###############################
# os.chdir(r'C:\Users\DELL\Desktop\personal\no_specs')
# list_files=[file for file in os.listdir() if file.endswith('txt') ] 
# # print(list_files)

# for file in list_files:
#      with open(file,'r+')as f:
#           line = f.read().split(' ')
#           # change 
#           line[0] ='0'
#           # print(line)
#           change =(' '.join(line))
#           # print(change)


# for file in list_files:
#      with open(file,'r+')as f:
#           ...
#           f.write(change)

########                                      ###############################
#           change text from file 
# os.chdir(r"C:\Users\DELL\Desktop\personal\specs")
# list_files=[file for file in os.listdir() if file.endswith('txt') and not file.startswith('classes') ] 
# # print(list_files)

# for file in list_files:
#      with open(file,'r+')as f:             # first read the file 
                                             
#           line = f.read().split('\n')                  # create a new line with changes
#           change= line[0].split(' ')
#           change[0] = '0' if change[0] == '15' else '1'
#           fchange =(' '.join(change))
#           # f.write(change)
         
# # for file in list_files:
#      with open(file,'r+')as f:
#           ...
#           print(fchange)
#           f.write(fchange)                  # perform


########################################################################
#                   flatten_dict


# def flatten_dict(dd, separator ='_', prefix =''):
#     return { prefix + separator + k if prefix else k : v
#              for kk, vv in dd.items()
#                     for k, v in flatten_dict(vv, separator, kk).items()
#              }           if isinstance(dd, dict) else { prefix : dd }

# initialising_dictionary
# ini_dict = {
#      'geeks1': {'Geeks1': {'for1': 7}},
#      'for2': {'geeks2': {'Geeks2': 3}},
#      'Geeks3': {'for3': {'for3': 1, 'geeks3': 4}}}

##################################################################################
#    EXTRACT PARTICULAR LINES FROM LIST      AS TXT FILE    from HTML file


# with open('part1.txt','r+') as f:
#      lines = f.readlines()
#      # print(lines.index('subhead1-r product-cell__description-name'))                   # what line we need to add index
#      # c=0
#      for num , line in enumerate(lines):
#           if 'subhead1-r product-cell__description-name' in line:                       # if that line exists
#                # c+=1
#                print(lines[num+1].split('data-test="product-cell-name">')[1].split('</h4>')[0])        # print and extract values from list 
#      # print(c)
##################################################################################
              
#        list_files=[file for file in os.listdir() if file.endswith('.jpg') and 'no' in file ] 
#          f.write(line.replace(''.join(line.split(' ')[0]) ,'0'))

#          Move files from one directory to another
# import shutil, os
# files = ['file1.txt', 'file2.txt', 'file3.txt']
# for f in files:
#     shutil.move(f, 'dest_folder')


'''  get specific files from a directory and move to another directory '''

# def get_all_jpg_files():
#     import os
#     os.chdir(r'C:\Users\Tushar\Desktop\Personal\personal\images\specs')
#     import glob
#     jpg_files = glob.glob('*.jpg')
# #     print(jpg_files) 
#     import shutil
#     # C:\Users\Tushar\Desktop\Personal\personal\all_images
#     for i in jpg_files:
#      #    os.rename(i,i.replace(' ','_'))
#      #    print(i)
#         shutil.move(i, r'C:\Users\Tushar\Desktop\Personal\personal\all_images')

# get_all_jpg_files()

##################################################################################
#         CONVERT FILE EXTENSION 
# from PIL import Image
# import os
  
# # importing the image 
# im = Image.open("geeksforgeeks.png")

  
# # converting to jpg
# rgb_im = im.convert("RGB")
  
# # exporting the image
# rgb_im.save("geeksforgeeks_jpg.jpg")


#         FOR MULTIPLE FILES
# from PIL import Image
# import os
# import os,shutil
# os.chdir(r'C:\Users\DELL\Desktop\JPEG I mages')
# old_l=os.listdir()
# for file in old_l:
#      name=file.split('.')[0]
#      im = Image.open(file)
#      rgb_im = im.convert("RGB")  
#      rgb_im.save(name+'.jpg')
#      os.remove(file)

################################################################

#                       EXTRACT TEXT AND READ IT FROM PDF FILES 
# import pyttsx3,PyPDF2

# book = open(r"D:\Work\statements\M1797 _Apr 2021.pdf", 'rb')
# pdfReader = PyPDF2.PdfFileReader(book)
# pages = pdfReader.numPages

# speaker = pyttsx3.init()
# # print(pages)
# for num in range(pages):
#      page = pdfReader.getPage(num)
#      text = page.extractText()
#      speaker.say(text)
#      speaker.runAndWait() 
#      #print(1,text)    

################################################################
#               NESTED LOOP PRACTICE

# real life example of nexted loop  CLOCK 
# for hour in range(1,25):
#      for minute in range(1,61):  # 60 sec > 1 min > 60 min ? 3600 hour 
#           for second in range(1, 61):
#                import time ; time.sleep(1)
#                print(second)
#           print()
#           print(f'{minute} minute done')
#           print()
#      print()
#      print(f'{hour} hour done')
#      print()


# matlab i ek SAATH SAATH badh , each j iteration stores i'th iteration  (1, 22, 333, 4444 ....)
#         for j in range(i) 
# here i will be keep on reducing  | 
# because range starts from i so j == (0,4)  , i == 1  j == (1,4)   ,i == 2 j ==(2,4)

# means NEGATIVE loop for inner iteration as it starts from (1,10) then j == (2,10) (3,10) ...
#         for j in range(i+1,10)             | for j in range(i-1)  == positive loop



##################################################################################

#                   Review of classes 
# class variables                = shared by all instances 
# class initilization (__init__) = multiple variables with different arguments  (multple children having different names or abilities)




################################################################## LATER  ##############################################################################


# loop while 

# def main():
#      students = [['Harry', 37.21], ['Berry', 37.21], ['Tina', 37.2], ['Akriti', 41], ['Harsh', 39]]
#      scores = ([ i[1] for i in students])
#      for item in scores:
#           if item > scores[0]:
#                # scores[0] = item
#                print(item)
#                item = scores[0] 
#                print(item)
#      # print(scores[0])
#      # print(item)
#      # print (scores)

# main()
# def main():
#      students = [['Harry', 37.21], ['Berry', 37.21], ['Tina', 37.2], ['Akriti', 41], ['Harsh', 39]]
#      print(sorted(students,key=lambda x:x[1]))

# main()

# assign 3 iterations from letters to each number in range(2,10)

# letters = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']

# def assign():
#      for i in range(len(letters)):
#           print(letters[i] ,end='\n'if i % 3 == 0 else ' ')
          
# assign()
################################################################## TRICKY ##########################################################

# send  unknown message on whatsapp
# API = 'https://api.whatsapp.com/send?phone=917984309091&text=hello'
# https://api.whatsapp.com/send?phone=+917814891872

############################################################################################################################
#                        Multiple Expressions in if else in one line 

# mark_attendance() if connect_fortclient() else print('\nLet me connect to fortclient first\t'),print('\nYour Connected to FortClient\nNow lets mark the attendance\n'),mark_attendance()
     


############################################################################################################################