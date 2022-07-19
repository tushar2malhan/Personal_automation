
import os
# from gtts import gTTS
import speech_recognition as sr


para = '''
Calculate the mean score of all users.                                            # sum([ i['age'] for i in Student.objects.all().values('age')  ])/2   
Find the user with the highest score.                                             # Student.objects.order_by('-age')[:1].values('age')[0]    
Find if there is any user with score less than 2 (return True/False)              #  User.objects.filter(age__lt=20).exists()           |  [True if Student.objects.filter(age__lt=20) else False]     | 
Find all the users which have score exactly equal to the mean score.              # Student.objects.filter(age=sum([ i['age'] for i in Student.objects.all().values('age') ])//2)
Change the score value of all users in a single query to 100 (one hundred)        #
Find all the users which don't have score 10 or 20                                # Student.objects.exclude(age__in=[10,20,31])
Print a list of all the score values from all users excluding the first 10 users. # Student.objects.exclude(id__in=[i for i in range(10)]).values('age')
'''

#                                               Questions
# d={}
# words=para.split()
# for i in words:
#      if i not in d:
#           d[i] =0
#      else:
#           d[i] += 1

# # print(d)


# shift all 0s to right hand side in [490,2,1,0,4,0,2,30,1,0]
# L=[1,2,4,5,0,6,0]
# def solve(L):
#      ''' Move all 0s to RHS '''
#      if len(L) == 0:
#           return L
#      k=0
#      for i in range(len(L)):
#           if L[i] !=0:
#                L[k] = L[i]
#                k+=1
#                print(L[i])
#      while k < len(L):
#             L[k] = 0
#             k+=1
    #  for o in range(k,len(L)):
    #       L[o] =0
#      return L 
# print(solve(L))

# using while loop and iter , pop items in marketdict
# markdict = [
#     {"Tom":67, "Tina": 54, "Akbar": 87, "Kane": 43, "Divya":73},
#     {"Tom2":672, "Tina2": 254, "Akbar2": 287, "Kane2": 243, "Divya2":273}     ]



# find groupings of number in string
# s='903jf90dj03jd'
# a = "abc12333abc3456abc7891"
# b = ([i.strip() for i in ''.join([i if i.isdigit() else ' ' for i in a ]).split(' ') if i])
# print(b)


# print(    sorted(d.items() , key = lambda kv:(kv[1],kv[0])) )



# d = {'a':0,'b':1,'c':0}

# if d['a'] >0:
#         print('a')
# elif d['b'] >0:
#         print('b')
# elif d['c'] >0:
#         print('c')
# elif d['d'] >0:
#         print('c')
# else:
#         print('not ok')


def check_if_else():
    ''' which are valid if else statements'''
    ...

    x,y=3,5
    if x<y : print('1') if x>y else print (0)

    if 4==1:print(0)
    elif 4==5:print(1)
    else:print(1)

    # Not Valid
    # if x<y : if x>y : print (0)
    # if x<y :  print (0) else:print(1)





# max value of an integer that can be set is infinite  |
#    in python 2 there was an internal limit but it was removed in python 3

# express hexadecimal value a5 as base-16 integer constant   >>> 0xA5
# print(int('0xA5',16))

# Express constant floating point value 3.2*10^-12  >>>  print(3.2e-12)
# print(3.2e12)                                     >>> 3200000000000.0


# Valid string literal                              >>> """foo'bar""" | "foo'bar" | 'foo\'bar'

# String literal for hexadecimal value 7E           >>> "x7E" 

# print(r'foo\\bar\nbaz') == foo\\bar\nbaz  >>>  because r'' ignores escape sequence


# a = [1,2,4,6]
# b = [1,2,3,4]

# for i in range(len(a)):
#     if a[i] not in b or b[i] not in a:
#         print(a[i])


# find second largest number in list
# ar = [3,21,4,5,2,1]
# min = ar[0]
# second_min = ar[0]

# for i in ar :
#     if i < min:
#         # second_max = max
#         # print(i)
#         min = i
#         print(min)

# print(min,second_max)
##################################################################################################################################################

''' Decorator is closuers for python 
where in the decorator > we call the function inside the sub function that needs to be run ( which is the decorated function )
same like func()   , we can pass any variable too in the sub functoin of the decorator and give values in the decorated function '''

# def hi(func):
#      print('hello this is the decorator ')
#      def wrapper(a):
#           print(f' calling {a} from inside the wrapper')
#           # return a+'hi'
#           return a   # this  refers to line 58 , where it gets called accordingly
#      return wrapper

# @hi
# def call(val):
#      print(val+'malhan')

# call('tushar')

# def input_name(func):
#      def wrapper(name):
#           name = input('Enter your name: ')
#           if name:
#                print(func(name))        # this will call decorated function,otherwise it has not meaning
#                return 'hi ' + name      
#           else:
#                return 'Please enter your name'
#      return wrapper



''' decorator   good example '''

# def fu(func):
#     print('\nhello')
#     return func

# @fu
# def  shell(funx):
#     def inner(a,c):
#         print('a')
#         if a==1:
#             return funx(12)
#         else:
#             return funx(c)
#     return inner

# @shell
# def f(z):
#     print('b')
#     return z+1

# print(f(4,12))

# f1 = shell(f)
# f1(1,12)


########################################################################################################################

# @input_name
# def say(name):
#      return 'hi'+name
# print(say('malhan'))

# from chatterbot import ChatBot
# from chatterbot.trainers import ListTrainer
# from chatterbot.trainers import CHatterBotCorpus Trainer

# logic_adapters = (['chatterbot.logic.Mathematical Evalution',
# 'chatterbot.logic.BestMatch'])


# smalltalk = ['hi','hai','how do you do',
# 'hello','hai','hey','hi','how are you','hey buddy','you are great','your great','hay','ok'
# ]


# talk1= ['best thing ever to take place ']

# list_trainer= List Trainer (my_bot)
# for item in (smalltalk,talk1):
#      list_trainer.train(item)


# corpus_trainer = ChatterBotCorpus Trainer(my_bot)
# corpus_trainer.train('chatterbot.corpus.english')

# print(my_bot.get_response('hello'))

# from gtts import gTTS
# voice = ''

# while True:
#      e = sr.Recognizer()
#      with sr.Microphone() as source:
#           try:
#                audio = e. listen(source)
#                text = e.recognize_google(audio)
#                print(text)
#                if text == "stop":
#                     break
#                text = e.recognize_google(audio)
#                voice = voice+str(text)

#           except:
#                 print("say something")

# we = gTTS(text=voice, Lang='en', slow=False)
# we.save("A.wav")

# from GoogleNews import GoogleNews
# googlenews = GoogleNews( )
# googlenews = GoogleNews(period='/7d' )
# googlenews.search( 'IND' )
# result=googlenews.result( )

# for x in result:
#      print("-"*50)
#      print("Title--", x['title'])
#      print("Date/Time--", x['date'])
#      print("Description--", x['desc'])

#      print("Link--", x['link'])


""" #############################################################################################################################################################################"""
#                                           AUTOMATION OF THE WEBPAGE


''' 
    Requests automation 
    with beautifulsoup  
'''

def requests_beautiful():
    
    ''' 
    get all TEXT , LINKS, IMAGES , TITLE  
    of the webpage
    from The base_url  '''
    import requests
    from bs4 import BeautifulSoup
    
    base_url = 'http://www.nytimes.com'
    r = requests.get(base_url)
    # print(r)


    # GET All the Text from the Page 
    # print(r.text)
    soup = BeautifulSoup(r.text,'html.parser')

    # GET title of the Page 
    print(soup.title.text)


    # GET all images from the page
    img_tags = soup.find_all('img')
    print( [img.get('src') for img in img_tags ])


    # GET all anchor tags from the page
    links = (soup.find_all('a'))                       
    [print(link.get('href')) for link in links ]       # GET href link
        

   # GET text from the class_ 
    for story_heading in soup.find_all(class_="indicate-hover css-vip0cf"): 
        print(story_heading.text)                     

<<<<<<< HEAD
# requests_beautiful()
=======
        if story_heading.a: 
            print(story_heading.a.text.replace("\n", " ").strip())
        else: 
            print(story_heading.contents[0].strip())
>>>>>>> 632aa1df7942dfcf34c9d616b24d02f8de79d875

# requests_beautiful()

''' selenium automation'''

from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook  ,load_workbook
wb = Workbook()  
url="https://www.mercadona.es/"
postalcode='08013'

def web_scrap_automation(url,postcode,*classnames):
     driver = webdriver.Chrome(executable_path=r"C:\Users\Tushar\Downloads\chromedriver.exe")
     driver.get(url)
     print('\t\t',driver.title)
     
     def get_values(sheet_num,sheet_name,driver,*innerclasses):     # here innerclasses refers to each element we sent when outer func is called 
               sheet = wb.create_sheet(sheet_num)
               sheet.title = sheet_name 
               
               #   get the Name
               sheet[f'A{1}']="PRODUCTS"
               name_tag= driver.find_elements_by_class_name(innerclasses[0])
               for name in range(len(name_tag)):
                    sheet[f'A{name+2}']=name_tag[name].text
                    print(name_tag[name].text)

               
               #   get the Price
               sheet[f'B{1}']="PRICES"
               price_tag=driver.find_elements_by_class_name(innerclasses[1])
               for price in range(len(price_tag)):
                    sheet[f'B{price+2}']= price_tag[price].text
                    print(price_tag[price].text)

               #  getting the image link
               images = driver.find_elements_by_xpath(innerclasses[2]) 
               sheet[f'C{1}']="IMAGES"
               for image in range(len(images)):
                    sheet[f'C{image+2}']=images[image].get_attribute('src')
                    print(images[image].get_attribute('src'))

               # get the weight
               weights =driver.find_elements_by_class_name(innerclasses[3]) 
               sheet[f'D{1}']="WEIGHT"
               for weight in range(len(weights)):
                    sheet[f'D{weight+2}'] = weights[weight].text
                    print(weights[weight].text)
               print()
                    
               print(f'\n\nDone , Data extracted and saved in your excel for the {sheet_name}')

     try:
          #              convert to english language
          l0=driver.find_element_by_xpath("/html/body/div[1]/nav/div[2]/div/i")
          l0.click()
          l1=driver.find_element_by_xpath("/html/body/div[1]/nav/div[2]/div/ul/li[6]/button")
          l1.click()
          
          search_bar = driver.find_element_by_xpath('//*[@id="root"]/header/div/div/form/div/input')
          
          #                   give the postcode
          try:
               search_bar.send_keys(postcode)      
               search_bar.send_keys(Keys.RETURN)
               l2=driver.find_element_by_xpath('//*[@id="root"]/header/div/div/form/input')
               l2.click()
               accept= driver.find_element_by_xpath('//*[@id="root"]/div[1]/div/div/button[2]')
               accept.click()
          except:
               print('No way to pass postal code on the webpage')

          time.sleep(10)
          #                        DYNAMIC FROM HERE 

          get_values('sheet_1','LANDING_PAGE',driver,*classnames)
          print('\n\n')
     
          # #                      View the products from inside 
          product_button = driver.find_element_by_class_name('banner-item__button')
          product_button.click()
          time.sleep(10)

          get_values('sheet_2','INNER_PAGE',driver,*classnames)
          print('\n\n')

          driver.close()      
        #       filename = fr"C:\Users\Tushar\Desktop\file_{postalcode}.xlsx"  
        #       os.mkdir(filename)                                                         
        #   wb.save(filename)                                                          # saving file to the excel 
        #   wb.remove(wb['Sheet'])
        #   wb.save(filename)  

     except Exception as e:
          print(f'Issue in the postal code ❌{postalcode}❌, check and verify it again')
          print(e)

# web_scrap_automation(url,postalcode,'product-cell__description-name','product-price__unit-price',"//div[@class='product-cell__image-wrapper']/img","product-price__extra-price",None)


######################################################################################################################################################

#    ubuntu multiple commands at once

# !wsl && cd /mnt/c/Users/DELL/Desktop/PYTHON/test/RealTimeObjectDetection/Tensorflow && git clone https://github.com/tensorflow/models

######################################################################################################################################################

     # Access variables from function

# global a : print(a)      
# create global variable  , use the same variable for the function , to call the global variable 
# when creating a function  => set this a inside function => func.a =1  => when func called func() , print(func.a)

# use it to create global and local variables inisde function
# globals()             # will return all global variables 
# x = globals()['a']    # will return specific global variable outside the function which is a


########################################################################
#                        find and change words in a list of strings via indexing

# ' '.join([ followers[i-1] for i,word in enumerate(followers) if word == 'Followers'])


########################################################################

#                    change text from file 

# os.chdir(r'C:\Users\DELL\Desktop\personal\no_specs')
# list_files=[file for file in os.listdir() if file.endswith('txt') ] 
# # print(list_files)

# for file in list_files:
#      with open(file,'r+')as f:             # first read the file 
                                             
#           line = f.read()                  # create a new line with changes
#           change=(line.replace(''.join(line.split(' ')[0]) ,'0'))

# for file in list_files:
#      with open(file,'r+')as f:
#           ...
#           f.write(change)                   # perform

########                                      ###############################
# os.chdir(r'C:\Users\DELL\Desktop\personal\no_specs')
# list_files=[file for file in os.listdir() if file.endswith('txt') ] 
# # print(list_files)

# for file in list_files:
#      with open(file,'r+')as f:
#           line = f.read().split(' ')
#           # change 
#           line[0] ='0'
#           # print(line)
#           change =(' '.join(line))
#           # print(change)


# for file in list_files:
#      with open(file,'r+')as f:
#           ...
#           f.write(change)

########                                      ###############################
#           change text from file 
# os.chdir(r"C:\Users\DELL\Desktop\personal\specs")
# list_files=[file for file in os.listdir() if file.endswith('txt') and not file.startswith('classes') ] 
# # print(list_files)

# for file in list_files:
#      with open(file,'r+')as f:             # first read the file 
                                             
#           line = f.read().split('\n')                  # create a new line with changes
#           change= line[0].split(' ')
#           change[0] = '0' if change[0] == '15' else '1'
#           fchange =(' '.join(change))
#           # f.write(change)
         
# # for file in list_files:
#      with open(file,'r+')as f:
#           ...
#           print(fchange)
#           f.write(fchange)                  # perform


########################################################################
#                   Pass List Iterables in startswith ()
#       print( 1 if     ask.startswith( ('7','8','9') )     else 0 )
########################################################################
#                   flatten_dict


# def flatten_dict(dd, separator ='_', prefix =''):
#     return { prefix + separator + k if prefix else k : v
#              for kk, vv in dd.items()
#                     for k, v in flatten_dict(vv, separator, kk).items()
#              }           if isinstance(dd, dict) else { prefix : dd }

# initialising_dictionary
# ini_dict = {
#      'geeks1': {'Geeks1': {'for1': 7}},
#      'for2': {'geeks2': {'Geeks2': 3}},
#      'Geeks3': {'for3': {'for3': 1, 'geeks3': 4}}}

##################################################################################
#    EXTRACT PARTICULAR LINES FROM LIST      AS TXT FILE    from HTML file


# with open('part1.txt','r+') as f:
#      lines = f.readlines()
#      # print(lines.index('subhead1-r product-cell__description-name'))                   # what line we need to add index
#      # c=0
#      for num , line in enumerate(lines):
#           if 'subhead1-r product-cell__description-name' in line:                       # if that line exists
#                # c+=1
#                print(lines[num+1].split('data-test="product-cell-name">')[1].split('</h4>')[0])        # print and extract values from list 
#      # print(c)
##################################################################################
              
#        list_files=[file for file in os.listdir() if file.endswith('.jpg') and 'no' in file ] 
#          f.write(line.replace(''.join(line.split(' ')[0]) ,'0'))

#          Move files from one directory to another
# import shutil, os
# files = ['file1.txt', 'file2.txt', 'file3.txt']
# for f in files:
#     shutil.move(f, 'dest_folder')


'''  get specific files from a directory and move to another directory '''

# def get_all_jpg_files():
#     import os
#     os.chdir(r'C:\Users\Tushar\Desktop\Personal\personal\images\specs')
#     import glob
#     jpg_files = glob.glob('*.jpg')
# #     print(jpg_files) 
#     import shutil
#     # C:\Users\Tushar\Desktop\Personal\personal\all_images
#     for i in jpg_files:
#      #    os.rename(i,i.replace(' ','_'))
#      #    print(i)
#         shutil.move(i, r'C:\Users\Tushar\Desktop\Personal\personal\all_images')

# get_all_jpg_files()

##################################################################################
#         CONVERT FILE EXTENSION 
# from PIL import Image
# import os
  
# # importing the image 
# im = Image.open("geeksforgeeks.png")

  
# # converting to jpg
# rgb_im = im.convert("RGB")
  
# # exporting the image
# rgb_im.save("geeksforgeeks_jpg.jpg")


#         FOR MULTIPLE FILES
# from PIL import Image
# import os
# import os,shutil
# os.chdir(r'C:\Users\DELL\Desktop\JPEG I mages')
# old_l=os.listdir()
# for file in old_l:
#      name=file.split('.')[0]
#      im = Image.open(file)
#      rgb_im = im.convert("RGB")  
#      rgb_im.save(name+'.jpg')
#      os.remove(file)

################################################################

#                       EXTRACT TEXT AND READ IT FROM PDF FILES 
# import pyttsx3,PyPDF2

# book = open(r"D:\Work\statements\M1797 _Apr 2021.pdf", 'rb')
# pdfReader = PyPDF2.PdfFileReader(book)
# pages = pdfReader.numPages

# speaker = pyttsx3.init()
# # print(pages)
# for num in range(pages):
#      page = pdfReader.getPage(num)
#      text = page.extractText()
#      speaker.say(text)
#      speaker.runAndWait() 
#      #print(1,text)    

################################################################
#               NESTED LOOP PRACTICE

# real life example of nexted loop  CLOCK 
# for hour in range(1,25):
#      for minute in range(1,61):  # 60 sec > 1 min > 60 min ? 3600 hour 
#           for second in range(1, 61):
#                import time ; time.sleep(1)
#                print(second)
#           print()
#           print(f'{minute} minute done')
#           print()
#      print()
#      print(f'{hour} hour done')
#      print()


# matlab i ek SAATH SAATH badh , each j iteration stores i'th iteration  (1, 22, 333, 4444 ....)
#         for j in range(i) 
# with i here j will be keep on Increasing  | 
# because range starts from i so j == (0,4)  , i == 1  j == (1,4)   ,i == 2 j ==(2,4)

# means NEGATIVE loop for inner iteration as it starts from (1,10) then j == (2,10) (3,10) ...
#         for j in range(i+1,10)             | for j in range(i-1)  == positive loop


##################################################################################

#                             Review of classes 


'''                        multiple classes inheritance               '''


# class variables                = shared by all instances  without using super().__init__()
# class initilization (__init__) = multiple variables with different arguments  (multple children having different names or abilities)

''' ways to call super class () '''
# A(self)  or A()         # called A ( parent class ) directly in B ( child class )
# A.__init__(self)        # here we dont need to inherit it explicitly , thats y we can call it directly  ie -> just class B
# A().__init__(self)      # no need to inherit class B(A)    #  init called twice
# super().__init__()      # only called when B class inherit in A   ie -> B(A)


class Parent1():

    Parent1 = "parent 1 "

    def __init__(self, name_1 ,*args,**kwargs) -> None:
        self.n = name_1 
        self.personality = 'patient' 
        self.args = args
        self.kwargs = kwargs
        # print('\t By default name_ is taken as ',name_1,'\n')

    def __str__(self) -> str:
        return f"\
         name is {self.n}\n\
         args is { self.args} \n\
         kwargs are { self.kwargs}\n\
         Personality is { self.personality}\
         "
    @staticmethod
    def static_method():
        print('\t this is a static method of parent 1')
         

# obj = Parent1('imaginary','i am ','arg of parent',name ='Gldy',age = 52)


class Parent2:
    ''' parent 2  '''

    def __init__(self,personality) -> None:
        self.personality = personality
    
    @staticmethod
    def static_method():
        print('\t this is a static method of parent 2')

    def __call__(self):
        print(f'\t used this to call for parent 2 parent which is {self.personality}')
    
        

class Child(Parent1,Parent2):

    child1 = 'child 1 '

    def __init__(self, name_1, *arg, **kwarg) -> None:
        self.name_2_bychild = name_1

        #    Copy Parent init with params  # mandatory 
        super().__init__(name_1, *arg, **kwarg)    
        #  - super() is used to access the parent class variable and methods  
        #    here init of parent class is overriden, so object will access this init variables 

        # self.personality = 'Aggresive'                    # override  # optional 
   
    @staticmethod
    def static_method():
        print('\t this is a static method of Child ')
    
    @classmethod
    def add_new__object(cls,string,string1):
        ''' Create new class variable & new obj from class method  '''
        cls.new_var = string1
        return cls(string)
    
    @classmethod
    def change_cls_variable(cls,value):
        print('\n\t Old value of Child class variable is ->',cls.child1,)
        print('\t Child class new variable given by child is ->',value)
        cls.child1 = value

    # @classmethod     >>>          # this class method only works for this class,not parents class 
    # so make sure if your changing parent class variable , dont use this class method on this func
    def change_parent_cls_variable(cls,value):
        print('\n\t Old value of Parent1 class variable is ->', Parent1.Parent1 )
        Parent1.Parent1 = value
        print('\t Parent1 class new variable given by child is ->',value)
        print('\t Look i confirmed it > ',Parent1.Parent1)
    
    def change_name_2(self,new_name):
        self.name_2_bychild = new_name
        print(f'\t name is changed to {new_name}')


obj2 = Child('tusharmalhan','i am ','arg of child',name ='Tushar',age = 23)

# obj3 = obj2.add_new__object('Obj_3','new_cls_var_2')
# print(obj3)
# print(obj3.new_var)
# print(Child.new_var)
# print(obj3.n)

# obj2()
# obj2.static_method()


''' child class can change both theirs class variable and parent class variable -(only by   Parent_class_name.class_variable = 'new cls var'  ) '''
# obj2.change_cls_variable('child_1 class variable  ')
# obj2.change_parent_cls_variable('parent_1 class variable  ')


'''                                 OVER-RIDING                           '''   
# Object looks First for instance variables  if not, then own class variable and then  search for inherited class variable
#  So if u want to access the parent class variable and methods , thats where we use super().__init__()

#  - if again variables name same "var2" , it will check position of super().__init__(), if called first then => B class instance Variable will over ride it
#    else if called after assigning "var2" value in child class  = it  will take variable from super ( PARENT ) class instance Variable
#           
#           # implicit call will call the parent class again and again [ in diamond shape class inheritance ] ,making it useless 
#           impilict call  ->  ClassName.__init__(self)
#           explicit call  ->  super().__init__()        or  super(ClassName,self).__init__()   == same 


'''      @property @getter @setter   and operator modulo  
        - we use [ property, getter, setter] > so that initilization attributes also change with the function
        - make sure variable name and function name are not same 
'''


class Trap():

    # private variable of class
    __private_class_variable = 'private class variable'
    # protected variable of class
    _protected_class_variable = 'protected class variable'

 

    def __init__(self,name) :
        self.name = name

    @property
    def email(self):
        ''' so whenever getter called we know about it '''
        print(f'Calling your email -> {self.name}@gmail.com')
        return f'{self.name}@gmail.com'


    @email.setter
    def email(self,new_email):
        ''' with setter we can change  other attributes too'''
        print('\nSetting your new  email ...')
        self.name = new_email
        print('Your new email is -> ',self.email)
  
    
    def __getitem__(self, index):
        ''' suppose  a is the object of class check and
        we want to get the value of a[0] as its a sequencial data type 
        we can use this method to get the value of a[0]
        else it cant subscribe to the object of class  check and throw error
        print(a[0])'''
        print('\n\t This will access your dict or string by the index ')
        return self.name[index]


    def __len__(self):
        ''' when u call object with len() method it will return the length of the object
        called from here 
        print(len(a))'''
        return len(['1',2,3])


    def __call__(self):
        ''' a() is  like a function '''
        print(self.name)

    
    def __add__(self, other):
        '''  
        where 2 objects a and b item and other gets added  
        print(a+b)     #   
        - This item will be taken from the constructor of the class
        '''
        return self.name + other.name
    
    
    def __mul__(self,other):
        return self.name +'\t'+ other.name * 2
    
    
    def __str__(self):
        ''' preference given first when object is printed as compared to __repr__
        print(object), print(self) to initate this '''
        return '\nThis class is for Practice , called by __str__ method'

   
    def __del__(self):
        ''' del object  to initate this method '''
        return ('Destructor called, Object deleted.')

print()
r = Trap('TusharMalhan')
# print(r.email)                # getter called here 
# r.email = 'malhan2'           # setter called here 

# print(r.email())              # getter called here 
# print(r.email_('malhan3'))    # setter called here 
# print(r.name)                 # without getter setter too , we can change attributes 

# print(r[0])                   # __getitem__()
# r()                           # __call__()
# print(len(r))                 # __len__()
# s = TrapArtist('Jimmy Malhan')
# print(r+s)                    # __add__()
# print(s)                      # __str__()
# print(r*s)                    # __mul__()
# print(r.__del__())            # __del__()

# print(r._protected_class_variable)                     # WE CAN CALL both PROTECTED variable of both class and instance >   DIRECTLY         
# print(r._TrapArtist__private_class_variable)           # for private , we need        object._classname__attribute  or getter method
# print(r.__dict__)                                      # call constructors params in dict format


''' 

PROTECTED == means we can call both class and instance protected variable   DIRECTLY   [ can call these attribute & methods  outside the class ]

PRIVATE ==       Even if variable or method is private or even if we access prvt attribute from parent class >>> 
|                we need  =  obj._classname__attribute  ==     CLASSNAME in between to call it  
|                else put private variable in public method and call it     DIRECTLY 

So Access private only by  :=   
|                        obj._classname__attribute   
|                        getter methods
|                        creating functions or private ones too                     [ we cannot call these attribute & methods  outside the class ]
'''

#                       Diamiond Shape Inheritance

class Class1:
    def m(self):
        print("In Class1")  
     
class Class2(Class1):
    def m(self):
        print("In Class2")
        # Class1.m(self)
        super().m()
 
class Class3(Class1):
    def m(self):
        print("In Class3")
        super().m()         
        # Class1.m(self)  
      
class Class4(Class2, Class3):
    '''  given attribute is first searched in the current class 
    if it’s not found then its searched in the parent classes.
    The order that is followed is known as a linearization 
    of the class Derived and this order is found out using 
    a set of rules called Method Resolution Order (MRO).'''
    def m(self):
        print("In Class4")  
        super().m()                 # explicit call which will do the mro and make the parent class call once 
        # Class2.m(self)
        # Class3.m(self)            # implicit call which will make the class 1 twice
      
# obj = Class4()
# obj.m()

from abc import ABCMeta,abstractmethod

class A(metaclass = ABCMeta):
    @abstractmethod
    def hello(self):
        print('\
            -This is an abstract method which ensures that\
             all the child classes must create a method name hello()\
             or else it will throw an error\
            - This works only with direct class inheritance not with sub class inheritance \
            - Thus no object cant be created for this class where abstract class method is set')

class child1(A):
    def hello(self):
        print('Hello from child 1 class')
    ...


# a = child1()
# a.hello()
    


################################################################## PROBLEMS  ##############################################################################


#           assign 3 iterations from letters to each number in range(2,10)

# letters = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']

# def assign():
#      for i in range(len(letters)):
#           print(letters[i],end='\n'if i % 3 == 0 and i !=0  else '  ')
          
# assign()
################################################################## TRICKY ##########################################################

# send  unknown message on whatsapp
# API = 'https://api.whatsapp.com/send?phone=917984309091&text=hello'
# https://api.whatsapp.com/send?phone=+917814891872

############################################################################################################################
#                        Multiple Expressions in if else in one line 

# mark_attendance() if connect_fortclient() else print('\nLet me connect to fortclient first\t'),print('\nYour Connected to FortClient\nNow lets mark the attendance\n'),mark_attendance()
     
#                        Check if name endswith with values given in list 

#  if  kingdom_name.endswith (tuple(i for i in [ 'a', 'e', 'i', 'o', 'u'] ))


############################################################################################################################
#                                                   GIT   HELP

#         Add git ignoring credentials file                                                                                                                   

# use gitignore to ignore the credentials file
# inside .gitignore > /dir/main_dir/          

# git add . 
# git commit -m "" 
# git push origin           >>> will ignore all  the directory or file

# SO WHEN U PUT FILE IN GITIGNORE , YOU WONT SEE THE FILE IN GIT STATUS   {  no relative path allowed , only absolute path }
# git rm --cached <foldername>   incase folder dosent show up in repository

# ** Suppose your changes made a HEAD branch and you lost the branch when checkouting to master
# -  git reflog                     # it shows all the HEADS and their changes , find ur hash id c526b1f
# - git checkout -b temp3 c526b1f  # will checkout to the hash id , so ur code will be updated based on that HEAD with new branch
# then you can merge the branch to master and push the changes 

############################################################################################################################
# git checkout -- filename.py      [ before git add . ]     >>> undo the latest changes you just did !     ( or do . to undo all changes in all files )
# git revert  hashcode             [ after git commit ]     >>> undo the changes and with the current changes in ur files , it will bring back old changes
#                                                              ( like with my new updated code in fortclient it brought back old README.MD file which was deleted )
#                     ** SENSITIVE COMMANDS **
# Q Want to  make your directory take back in time  with that git commit ,so all the current commits and files will be lost  permanently
# git reset --hard hashcode        [ after git commit ]     

# Q Want any old commit back from log ?
# git checkout  hashcode              
###########################################################################################################################

#                                            CSV basic operations


# from csv import writer
# with open('posts.csv', 'w') as csv_file:
#     csv_writer = writer(csv_file)
#     headers = ['Title', 'Link', 'Date']
#     list_data=['03','Smith','Science']
#     list_of_dicts=[
#         {'ID':1,'NAME':'William','RANK':5532,'ARTICLE':1,'COUNTRY':'UAE'},
#         {'ID':6,'NAME':'William','RANK':5532,'ARTICLE':1,'COUNTRY':'UAE'}
#         ]
    # a=([(k,v) for k,v in dict.items()])  
          

    #        - This will set the headers
    # csv_writer.writerow(headers)   
    #      
    #        - This will add the data
    # csv_writer.writerow(list_data)         
    #      
    #       - If you comment and use writerow
    #       - it will Write new data to the file
    # csv_writer.writerow(list_data[0])
    #       
    #       - List of tuples will be entered in each row 
    #       - Do the loop and write the row automatically!
    # keys = ([  keys for keys in list_of_dicts[0] ])
    # print(keys)
    # csv_writer.writerow(keys)

    #       - ADD multiple values to the file
    # for values in list_of_dicts:
    #     csv_writer.writerow(values.values())



    #       - OR
    # from csv import DictWriter

    #       - Field name needs to be matched
    # field_names = ['ID','NAME','RANK','ARTICLE','COUNTRY']
    # dict_object = DictWriter(csv_file, fieldnames=field_names)

    #       - This will write the headers, no need to pass anything
    #       - will pick it from dict_object['fieldnames']
    # dict_object.writeheader()
    # dict_object.writerow(dict)
############################################################################################################################
#                                                Excel file basic operations


'''
POINTS 
 - IF FILE IS OPENED , THEN IT WILL NOT SAVE THE FILE , ERROR 
 - For each specific column , seperate for loop will run , which make sures to get data for each 
 specific column first.
'''

#                               Save  data in excel file even if file does not exists

# from openpyxl import Workbook

#                                Can work with multiple sheets at Once
# wb = Workbook()
# wb['Sheet'].title = "Report of Automation"               # 1 way to change title of sheet
# sh1 = wb.active                                        # active or create_sheet       # same same                 
# sh1['A1'].value = 'HEAD 0'
# sh1['B1'].value = 'HEAD 1'
# sh1['C1'].value = 'HEAD 2'

# sh1['A2'].value = 'VAL 1'
# sh1['B2'].value = 'VAL 2'

# for a in range(5):
#     sh1['A'+str(a+1)].value = a

# ws2 = wb.create_sheet()
# ws2.title = "5>!"                       # This is the title of the sheet ,   # 2 way to change title of sheet
# ws2['A1'] = '6'
# ws2['B1'] = '7'
# ws2['C1'] = '8'

# ws3 = wb.create_sheet()                 # If empty then sheet name is sheet by (default)

# list_of_dicts=[
#         {'ID':1,'NAME':'William','RANK':5532,'ARTICLE':1,'COUNTRY':'UAE'},
#         {'ID':6,'NAME':'William','RANK':5532,'ARTICLE':1,'COUNTRY':'UAE'},
#         {'ID':2,'NAME':'William','RANK':5532,'ARTICLE':3,'COUNTRY':'UAE'}
#         ]

# sh3 = wb.create_sheet()
# sh3.title = "sheet3"

# append keys and values in excel file 
# sh3.append([i for i in list_of_dicts[0].keys()])   # KEYS OF THE DICT
# for values in list_of_dicts:
#     sh3.append([i for i in values.values()])       # VALUES OF THE DICT

# wb.save('filename.xlsx')                           # MAKE SURE YOU save the file 



###################################################################

# Add and install Older version of Python 
# Remove new vesion of python from system variables 
# check version , old will be  placed
############################################################################################################################
#                                                Convert py to exe 
# cd into dir pyinstaller (pip install pyinstaller) > pyinstaller filename.py
# cd into dist > then run filename.exe , thus can send others ur software
# (FIle outside dist dir , runs the file without terminal so for bussiness use)

# HAVE A SINGLE FILE AS EXE , INSTEAD OF MULTIPLE DIRECTORIES FOR THE SOFTWARE
# cd into dist > then run filename.exe , thus can send others ur software
# DONT CONVERT IMG TO ICO file as icon for ur softw
# pyinstaller  --onefile  --name = test_fortclient c:/Users/Tushar/Desktop/python/completed_automation/connect_fortclient.py 


############################################################################################################################

#                                        Selenium easy ways  
# find element by text     =>  driver.find_element_by_xpath("//*[contains(text(), 'Call +XX XXXXXXXX72‎')]")
# find all links in table  =>  elements = driver.find_elements(By.XPATH,"//table[@class = 'table']//td/a")

############################################################################################################################

#                                           DEBUGGER TOOLS 

# import pdb;pdb.set_trace()
# pip install ipdb     |   python -m ipdb module.py         |  s > next step ,  c > continue   , q > quit 

############################################################################################################################

#                                          new vs init     
# class UppercaseTuple(tuple):

#     def __init__(self, list) -> None:
#         print(f"Start changes for {list}")

#         for i, item in enumerate(list):
#             self[i] = item.upper()

# print(UppercaseTuple(["my list", "uppercased tuple"]))



